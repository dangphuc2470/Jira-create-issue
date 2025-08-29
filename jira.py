import openpyxl
import urllib.parse
import json
import os
from datetime import datetime
import calendar

# --- Configuration ---
INPUT_EXCEL_FILE = "input_data.xlsx"
OUTPUT_EXCEL_FILE = "input_data.xlsx"
SHEET_NAME = "Sheet1" # The name of the sheet in your Excel file containing the data
SESSION_CONFIG_FILE = "session_config.json"

def load_session_config():
    """Load session configuration from JSON file."""
    try:
        with open(SESSION_CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"Error: {SESSION_CONFIG_FILE} not found. Please create the session config file.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in {SESSION_CONFIG_FILE}: {e}")
        return None

def build_curl_template(config):
    """Build the curl command template from config."""
    if not config:
        return None, None
    
    # Build headers
    header_parts = []
    for key, value in config['headers'].items():
        header_parts.append(f"--header '{key}: {value}'")
    
    # Build the curl template
    header_lines = ' \\\n'.join(header_parts)
    curl_template = f"""curl --location --request POST '{config['base_url']}' \\
{header_lines} \\
--header 'Cookie: {config['cookies']}' \\
--data-raw '{{DATA_PAYLOAD}}' \\
--insecure"""
    
    return curl_template, config['data_payload_template']

def generate_curl_command_for_jira(start_date_input, end_date_input, summary_text, actual_date_input):
    """
    Constructs a full curl command string for Jira API call using session config.

    Args:
        start_date_input (str or datetime.datetime): The start date, expected in DD/MM/YYYY if string.
        end_date_input (str or datetime.datetime): The end date, expected in DD/MM/YYYY if string.
        summary_text (str): The summary text for the Jira issue.
        actual_date_input (str or datetime.datetime): The actual date, expected in DD/MM/YYYY if string.

    Returns:
        str: The complete curl command string, or None if a date format error occurs.
    """
    # Load session configuration
    config = load_session_config()
    if not config:
        return None
    
    curl_template, data_payload_template = build_curl_template(config)
    if not curl_template:
        return None
    
    try:
        # --- Step 1: Normalize input dates to string DD/MM/YYYY if they are datetime objects ---
        if isinstance(start_date_input, datetime):
            start_date_str = start_date_input.strftime("%d/%m/%Y")
        else:
            start_date_str = str(start_date_input)

        if isinstance(end_date_input, datetime):
            end_date_str = end_date_input.strftime("%d/%m/%Y")
        else:
            end_date_str = str(end_date_input)

        if isinstance(actual_date_input, datetime):
            actual_date_str = actual_date_input.strftime("%d/%m/%Y")
        else:
            actual_date_str = str(actual_date_input)

        # --- Step 2: Parse the DD/MM/YYYY input strings into datetime objects ---
        start_date_obj = datetime.strptime(start_date_str, "%d/%m/%Y")
        end_date_obj = datetime.strptime(end_date_str, "%d/%m/%Y")
        actual_date_obj = datetime.strptime(actual_date_str, "%d/%m/%Y")

        # --- Step 2.5: Calculate end of month for the end date ---
        # Get the last day of the month for the end date
        last_day_of_month = calendar.monthrange(end_date_obj.year, end_date_obj.month)[1]
        end_date_obj = end_date_obj.replace(day=last_day_of_month)

        # --- Step 3: Format the datetime objects into Jira's required "d/MM/yy" format ---
        # Use %d for day without leading zero, %m for month with leading zero, %y for 2-digit year
        formatted_start_date = start_date_obj.strftime("%d/%m/%y").lstrip("0")
        formatted_end_date = end_date_obj.strftime("%d/%m/%y").lstrip("0")
        formatted_actual_date = actual_date_obj.strftime("%d/%m/%y").lstrip("0")

        # --- Step 4: Calculate remaining estimate in days ---
        # Calculate the difference between actual date and start date
        time_difference = actual_date_obj - start_date_obj
        remaining_estimate_days = max(0, time_difference.days)  # Ensure it's not negative

    except ValueError as e:
        print(f"Date format error: {e}. Please ensure dates in Excel are in DD/MM/YYYY format.")
        return None
    except TypeError as e:
        print(f"Data type error for dates: {e}. Ensure cells contain valid date strings or datetime objects.")
        return None

    # --- Step 4: URL encode all dynamic values ---
    encoded_summary = urllib.parse.quote(summary_text, safe='')
    encoded_start_date = urllib.parse.quote(formatted_start_date, safe='')
    encoded_end_date = urllib.parse.quote(formatted_end_date, safe='')
    encoded_actual_date = urllib.parse.quote(formatted_actual_date, safe='')

    # --- Step 5: Populate the data payload template ---
    populated_data_payload = data_payload_template.format(
        ATL_TOKEN=config['atl_token'],
        FORM_TOKEN=config['form_token'],
        SUMMARY=encoded_summary,
        START_DATE=encoded_start_date,
        END_DATE=encoded_end_date,
        ACTUAL_DATE=encoded_actual_date,
        REMAINING_ESTIMATE=remaining_estimate_days
    )

    # --- Step 6: Insert the populated data payload into the base curl template ---
    full_curl_command = curl_template.replace(
        "{DATA_PAYLOAD}",
        populated_data_payload
    )

    return full_curl_command

# --- Main Script Execution ---
def main():
    """
    Reads data from Excel, generates curl commands, and writes them back.
    """
    print(f"Attempting to read from '{INPUT_EXCEL_FILE}'...")
    try:
        # Load the workbook and select the active sheet
        wb = openpyxl.load_workbook(INPUT_EXCEL_FILE)
        ws = wb[SHEET_NAME]

        # Add a header for the Curl Command column if it doesn't exist.
        # Assuming Curl Command will be in Column E (5th column) if A,B,C,D are data.
        curl_column_index = 5 # E is the 5th column
        if ws.cell(row=1, column=curl_column_index).value != "Generated Curl Command":
            ws.cell(row=1, column=curl_column_index).value = "Generated Curl Command"

        # Iterate through rows starting from the second row (assuming row 1 has headers)
        # Columns: A=Start Date, B=End Date, C=Summary
        # Actual Date will be set to the same as End Date
        for row_idx in range(2, ws.max_row + 1):
            start_date_cell = ws.cell(row=row_idx, column=1) # Column A
            end_date_cell = ws.cell(row=row_idx, column=2)   # Column B
            summary_cell = ws.cell(row=row_idx, column=3)    # Column C

            start_date = start_date_cell.value
            end_date = end_date_cell.value
            summary = summary_cell.value
            # Set actual_date to be the same as end_date
            actual_date = end_date

            # Add [Mobile] to the summary
            summary = f"[Mobile] {summary}"

            # Check if all required data for the row is present
            if start_date is None or end_date is None or summary is None:
                ws.cell(row=row_idx, column=curl_column_index).value = "Missing data in row."
                print(f"Skipping row {row_idx}: Missing data.")
                continue

            # Generate the curl command
            curl_command = generate_curl_command_for_jira(
                start_date,
                end_date,
                summary,
                actual_date
            )

            # Write the generated curl command back to the Excel sheet
            if curl_command:
                ws.cell(row=row_idx, column=curl_column_index).value = curl_command
            else:
                ws.cell(row=row_idx, column=curl_column_index).value = "Error generating command (check console)."

        # Save the modified workbook to a new file
        wb.save(OUTPUT_EXCEL_FILE)
        print(f"\nCurl commands generated successfully and saved to '{OUTPUT_EXCEL_FILE}'")

    except FileNotFoundError:
        print(f"Error: The input file '{INPUT_EXCEL_FILE}' was not found.")
        print("Please ensure 'input_data.xlsx' is in the same directory as the script.")
    except KeyError:
        print(f"Error: Sheet '{SHEET_NAME}' not found in '{INPUT_EXCEL_FILE}'.")
        print("Please check the sheet name in your Excel file or update SHEET_NAME variable.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()
